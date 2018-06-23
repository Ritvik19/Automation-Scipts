# xl2txt source_file_path destination_folder

import openpyxl, sys, os
# get the sorce file path
xlpath = sys.argv[1]
# get the name of the source file
filename = os.path.basename(xlpath)
# get the destination folder
txtpath = sys.argv[2]
# array to store file objects for the varios work sheets
txtfiles = []

# initialising workbook object
workbook_obj = openpyxl.load_workbook(xlpath)
# getting the list of sheets
sheetlist = workbook_obj.get_sheet_names()

# for each sheet
for sheet_name in sheetlist:
    # initialising worksheet object
    worksheet_obj = workbook_obj.get_sheet_by_name(sheet_name)
    # initialising blank data string
    datastring = ""
    # for each row in work sheet
    for i in range(1, worksheet_obj.max_row+1):
        # for each column in row
        for j in range(1, worksheet_obj.max_column+1):
            # getting the data and adding inthe data string
            data = str(worksheet_obj.cell(row = i, column = j).value)
            if data != "None":
                datastring += data + "\t"
            else:
                datastring += "\t" + "\t"
        datastring += "\n"
    # append text file object
    # file name = destination_folder+soruce_file_name+_+sheet_name+.txt
    txtfiles.append(open(txtpath+filename+"_"+sheet_name+".txt", "w"))
    # write the datastring
    txtfiles[-1].write(datastring)
    # close the text file
    txtfiles[-1].close()

# close the workbook
workbook_obj.close()
